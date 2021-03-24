import pandas as pd
import os, re, hashlib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from get_output_from_NuSMV import*

#エクセルファイルへの出力するクラス
class CreateResultFile:
    wb = Workbook()
    ws = wb.active
    ws.title = "temp"

    FILL_TRUE= PatternFill(patternType='solid', fgColor='B0E0E6')
    FILL_FALSE= PatternFill(patternType='solid', fgColor='FFE4B5')
    excel_file_name = "couterexample_data.xlsx"
    @classmethod
    def save(cls):
        cls.wb.save(CreateResultFile.excel_file_name)

    def __init__(self, loop_start_pos):
        self.loop_start_pos = loop_start_pos
        self.row_start_idx=2
        self.col_start_idx = 0
#エクセルファイルへの書き込み処理
    def write_to_excel(self, title, li, data_type):
        self.col_start_idx+=1
        data = pd.DataFrame({title:li})
        rows = dataframe_to_rows(data, index=False, header=True)
        border=Border(bottom=Side(style="dotted" , color = "ff0000"))

        for row_no, row in enumerate(rows, self.row_start_idx):
           
            for col_no, value in enumerate(row, self.col_start_idx):

                self.ws.cell(row=row_no, column=col_no, value=value) # 1セルづつ書込む
                if row_no == self.loop_start_pos:
                    self.ws.cell(row=row_no+1, column=col_no).border = border
                if data_type == "spec":
                    if value == "True":
                        self.paint_cells(col_no,row_no,self.FILL_TRUE)
                        #self.ws[self.num2alpha(col_no)+str(row_no)].fill = self.FILL_TRUE
                    elif value == "False":
                        self.paint_cells(col_no,row_no,self.FILL_FALSE)
                        #self.ws[self.num2alpha(col_no)+str(row_no)].fill= self.FILL_FALSE
                elif type(data_type) == list:
                    if row_no != 1 and row_no-2 < len(data_type) and data_type[row_no-2] == True:
                        self.paint_cells(col_no,row_no,self.FILL_TRUE)
                    
        self.ws.column_dimensions[self.num2alpha(col_no)].width = 15
    # セルを色付けする関数
    def paint_cells(self,  col_no, row_no, fill):
        self.ws[self.num2alpha(col_no)+str(row_no)].fill = fill     
    #セルへの書き込みする関数
    def write_to_cell(self, col_no, row_no, value):
        self.ws.cell(row=row_no, column=col_no, value=value)
    #数字をアルファベット変換する関数
    def num2alpha(self,num):
        if num<=26:
            return chr(64+num)
        elif num%26==0:
            return self.num2alpha(num//26-1)+chr(90)
        else:
            return self.num2alpha(num//26)+chr(64+num%26)
#エクセルファイルへの書き込みを順次呼び出す関数
class OutputTabular:
    def __init__(self, result, num):
        self.result = result
        self.non_space_spec = result["specification"].replace(" ", "").encode()
        self.spec_hash = str(int(hashlib.sha1(self.non_space_spec).hexdigest(), 16) % (10**8)).zfill(8)
        if "loop_start_pos" in self.result.keys() and len(self.result["loop_start_pos"]) != 0:
            self.loop_start_pos = self.result["loop_start_pos"][0]
        else:
            self.loop_start_pos = None
        
        CreateResultFile.ws= CreateResultFile.wb.create_sheet("counterexample_"+self.spec_hash)
        
        self.excel = CreateResultFile(self.loop_start_pos)
        self.excel.write_to_cell(1, 1, self.result["specification"])
        self.excel.write_to_excel("state_num", [str(i+1) for i in range(len(self.result["state_num"]))], "state")
        self.output()
        CreateResultFile.save()
        
        
    def output(self):
        temp_dict = {i:[] for i in self.result["state_variables"][0].keys()}
        for state in self.result["state_variables"]:
            for key in  temp_dict.keys():
                temp_dict[key].append(state[key])
        for key in temp_dict.keys():
            self.excel.write_to_excel(key, temp_dict[key], "variable")

#検査式などと反例モデルを用いて検査
class Trace:
    def __init__(self, model_data, formulas, result, num, formula_file):
        self.formula_file = formula_file
        self.num = num
        self.model_data = model_data
        self.result = result
        self.non_space_spec = result["specification"].replace(" ", "").encode()
        self.spec_hash = str(int(hashlib.sha1(self.non_space_spec).hexdigest(), 16) % (10**8)).zfill(8)
        if CreateResultFile.ws.title != "temp":
            CreateResultFile.ws=CreateResultFile.wb.create_sheet(self.spec_hash)
        else:
            CreateResultFile.ws.title=str(self.spec_hash)
        
        self.excel = CreateResultFile(model_data.loop_start_pos)
        self.excel.write_to_cell(1, 1, self.result["specification"])
        #データ書き込み
        self.max_state = self.model_data.max_state
        self.excel.write_to_excel("state_num", [str(i+1) for i in range(self.max_state)], "state")
        self.path = "verification_counterexample.smv"
        #self.init_next_state = "init(next_state) := 2;"
        #self.state_num = 2
        if formulas == None:
            self.formula_file = "formulas.txt"
            self.formula_li = self.get_formula_li(self.formula_file)
        elif len(formulas) >0 and formulas[0] in "SPEC":
            self.formula_li = formulas
        else:

            self.formula_li = self.create_formula_li(formulas)
            self.formula_li = sorted(set(self.formula_li), key = self.formula_li.index)
            
        self.trace()

        CreateResultFile.save()

    #検査に用いる検査式，条件式，変数名をリストに整理
    def create_formula_li(self, formulas):
        li = []
        skip_tf = False
        if len(formulas) ==0 or type(formulas[0]) is str:
            return formulas
        for index, item in enumerate(formulas): 
            if skip_tf == True:
                skip_tf = False
                continue
            ltl_tf = False
            temp_formula = "( " + item.formula
            temp_ltl_ope_li = list(dict.fromkeys(item.ltl_ope_li))
            for i in reversed(temp_ltl_ope_li):
                if i == "G" or i=="F" or i=="X":
                    ltl_tf = True
                if  "G" in i or "F" in i or "X" in i :
                    temp_formula = i + " " + temp_formula
            
            if item.ltl_ope_li[-1] == "U" and index+1 <= len(formulas) :
                skip_tf = True
                temp_formula = temp_formula + " U " + formulas[index+1].formula +  " )"
            else:
                temp_formula +=  " )"
            temp_formula =  "SPEC " + temp_formula
            if ltl_tf:
                temp_formula = "LTL" + temp_formula 
            li.append(temp_formula)
            

        return li
    #検査式，条件式，変数名のリストと反例モデルを用いて検査
    def trace(self):
        for formula in self.formula_li:
            result_li = []
            if "SPEC" in formula:
                
                for i in range(self.max_state):
                    self.state_num = i+1
                    self.create_smv_file(formula)
            
                    output = Get_Output_From_NuSMV(self.path)
                    result = output.result
                    if result[0] == "Error":
                        break
                    elif result[0] == True:
                        result_li.append("True")
                    else:
                        result_li.append("False")
                else:
                    self.excel.write_to_excel(formula, result_li, "spec")

            else:
                if "=" in formula or "<" in formula or ">" in formula:
                    variable = ""
                    symbol = ""
                    value = ""
                    symbol_tf = True
                    for i in formula:
                        if i.isalpha() or i.isdigit() or i in [".","_","$","#","-", "[", "]"]:
                            if symbol_tf:
                                variable +=i
                            else:
                                value += i
                        elif i == " ":
                            continue
                        else:
                            symbol_tf = False
                            symbol+=i
                    if variable not in self.result["state_variables"][0].keys():
                        continue
                    formula_tf_li = self.formula_tf(variable, symbol, value)
                    if variable in self.result["state_variables"][0].keys():
                        value_li = []
                        for state in self.result["state_variables"]:
                            value_li.append(state[variable])
                        self.excel.write_to_excel(variable, value_li, formula_tf_li)
                else:
                    if formula in self.result["state_variables"][0].keys():
                        value_li = []
                        for state in self.result["state_variables"]:
                            value_li.append(state[formula])
                        self.excel.write_to_excel(formula, value_li, "variable")

                    
                #self.excel.write_to_excel(formula, result_li, "None")


    #条件式の検査
    def formula_tf(self, variable, symbol, value):
        formula_tf_li = [False for _ in range(self.max_state)]
        for i, state_dict in enumerate(self.result["state_variables"]):
            if symbol == "!=":                
                if state_dict[variable] != value:
                    formula_tf_li[i] = True
                continue
            elif "=" in symbol:
                if state_dict[variable] == value:
                    formula_tf_li[i] = True
                    continue
            if "<" in symbol:
                if int(state_dict[variable]) < int(value):
                    formula_tf_li[i] = True
            elif ">" in symbol:
                if int(state_dict[variable]) > int(value):
                    formula_tf_li[i] = True
        
        return formula_tf_li
    #適宜初期値を変えた反例モデルを生成
    def create_smv_file(self, formula):
        self.model_data.formula = formula.replace(".", "_")
        self.model_data.assign = self.model_data.create_ASSIGN(self.state_num)
        self.model_data.spec = self.model_data.formula
        self.model_data.write_counterexample_model(self.path)
       
    #テキストファイルから検査式などを取得
    def get_formula_li(self, filename):
        with open(filename) as f:
            li = f.readlines()
            li = [re.sub("[\r\n]+$",'', i )for i in li]
            temp =[]
            for i in li:
                if "SPEC" in i:
                    temp.append(i)
                else:
                    temp.append(i.replace(" ", ""))

            li=temp
            li = [i for i in li if i != "" ]
            return li
